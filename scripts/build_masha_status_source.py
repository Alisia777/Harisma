from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

FIELD_ALIASES = {
    'statusDate': ['status_date', 'дата', 'дата статуса'],
    'marketplace': ['marketplace', 'площадка', 'mp'],
    'articleKey': ['article_key', 'articlekey', 'sku_code', 'sku', 'артикул', 'код sku'],
    'article': ['article', 'артикул wb'],
    'skuName': ['sku_name', 'name', 'товар', 'название'],
    'ownerName': ['owner_name', 'owner', 'ответственный', 'менеджер'],
    'ownerLogin': ['owner_login', 'login'],
    'statusCode': ['status_code', 'статус код', 'status'],
    'statusText': ['status_text', 'статус', 'статус текст'],
    'riskFlag': ['risk_flag', 'risk', 'риск'],
    'riskLevel': ['risk_level', 'уровень риска'],
    'nextStep': ['next_step', 'следующий шаг', 'что делаем'],
    'comment': ['comment', 'комментарий'],
    'dueDate': ['due_date', 'срок', 'дедлайн'],
    'needsTask': ['needs_task', 'создать задачу', 'нужна задача'],
    'priority': ['priority', 'приоритет'],
    'rootCause': ['root_cause', 'причина', 'root cause'],
    'meetingType': ['meeting_type', 'встреча'],
    'linkedCluster': ['linked_cluster', 'кластер'],
    'linkedWarehouse': ['linked_warehouse', 'склад'],
    'linkedDocUrl': ['linked_doc_url', 'ссылка на документ'],
    'sourceName': ['source_name', 'источник'],
    'sourceType': ['source_type', 'тип источника'],
    'updatedBy': ['updated_by', 'обновил']
}


def norm_header(value: str) -> str:
    return ''.join(ch.lower() for ch in str(value).strip() if ch.isalnum() or ch in {'_', ' '}).replace(' ', '_')


def map_headers(headers: List[str]) -> Dict[str, int]:
    normalized = [norm_header(h) for h in headers]
    mapping: Dict[str, int] = {}
    for target, aliases in FIELD_ALIASES.items():
        alias_set = {norm_header(a) for a in aliases + [target]}
        for idx, header in enumerate(normalized):
            if header in alias_set:
                mapping[target] = idx
                break
    return mapping


def read_csv(path: Path) -> List[dict]:
    with path.open('r', encoding='utf-8-sig', newline='') as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return []
    headers = rows[0]
    mapping = map_headers(headers)
    payload = []
    for row in rows[1:]:
        if not any(str(cell).strip() for cell in row):
            continue
        item = {field: (row[idx] if idx < len(row) else '') for field, idx in mapping.items()}
        payload.append(item)
    return payload


def read_xlsx(path: Path) -> List[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or '') for cell in rows[0]]
    mapping = map_headers(headers)
    payload = []
    for row in rows[1:]:
        if not any(cell not in (None, '') for cell in row):
            continue
        item = {field: (row[idx] if idx < len(row) and row[idx] is not None else '') for field, idx in mapping.items()}
        payload.append(item)
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description='Build portal-ready JSON for Masha status source.')
    parser.add_argument('--input', required=True, help='Path to source CSV/XLSX file')
    parser.add_argument('--output', required=True, help='Path to output JSON file')
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        raise SystemExit(f'Input file not found: {src}')

    if src.suffix.lower() == '.csv':
        payload = read_csv(src)
    elif src.suffix.lower() in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        payload = read_xlsx(src)
    else:
        raise SystemExit('Supported formats: CSV, XLSX')

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Wrote {len(payload)} rows to {out}')


if __name__ == '__main__':
    main()
