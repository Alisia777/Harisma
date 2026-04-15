import json
import math
import sys
import datetime as dt
from pathlib import Path

import openpyxl


def num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return None
        except TypeError:
            pass
        return float(value)
    try:
        parsed = float(str(value).replace(',', '.'))
        if math.isnan(parsed):
            return None
        return parsed
    except Exception:
        return None


def clean(value):
    if value is None:
        return ''
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    return '' if text in {'None', 'nan'} else text


def build_payload(xlsx_path: Path, prices_json: Path | None = None, skus_json: Path | None = None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Сводная']

    price_map = {}
    if prices_json and prices_json.exists():
        prices = json.loads(prices_json.read_text(encoding='utf-8'))
        for market in ['wb', 'ozon', 'ym']:
            for row in prices.get('platforms', {}).get(market, {}).get('rows', []):
                price_map[(market, row.get('articleKey'))] = row

    sku_map = {}
    if skus_json and skus_json.exists():
        skus = json.loads(skus_json.read_text(encoding='utf-8'))
        sku_map = {row.get('articleKey'): row for row in skus}

    rows = {'wb': [], 'ozon': [], 'ym': []}
    for r in range(17, ws.max_row + 1):
        article_key = clean(ws.cell(r, 6).value)
        brand = clean(ws.cell(r, 7).value)
        if not article_key or brand != 'Алтея':
            continue
        sku = sku_map.get(article_key, {})
        base = {
            'articleKey': article_key,
            'article': sku.get('article', article_key),
            'name': sku.get('name', ''),
            'owner': (sku.get('owner') or {}).get('name', ''),
            'brand': brand,
            'strategy': clean(ws.cell(r, 3).value),
            'strategyNote': clean(ws.cell(r, 4).value),
            'direction': clean(ws.cell(r, 5).value),
            'cost': num(ws.cell(r, 8).value),
            'productStatus': clean(ws.cell(r, 9).value),
            'wbBlock': clean(ws.cell(r, 10).value),
            'trafficSignal': clean(ws.cell(r, 11).value),
            'stockWb': num(ws.cell(r, 12).value),
            'stockOzon': num(ws.cell(r, 13).value),
            'stockWarehouse': num(ws.cell(r, 14).value),
            'sales7Wb': num(ws.cell(r, 16).value),
            'sales7Ozon': num(ws.cell(r, 17).value),
            'sales7Total': num(ws.cell(r, 18).value),
            'turnoverWbDays': num(ws.cell(r, 20).value),
            'turnoverOzonDays': num(ws.cell(r, 21).value),
            'turnoverTotalDays': num(ws.cell(r, 23).value),
            'arrivalDate': clean(ws.cell(r, 24).value),
            'firstPrice': num(ws.cell(r, 52).value),
            'discountFromPricePct': num(ws.cell(r, 53).value),
            'recommendedFirstPrice': num(ws.cell(r, 55).value),
            'seedReason': clean(ws.cell(r, 67).value),
        }
        market_defs = {
            'wb': {
                'currentFillPrice': num(ws.cell(r, 57).value),
                'currentClientPrice': num(ws.cell(r, 60).value),
                'currentSppPct': num(ws.cell(r, 64).value),
                'profitabilityPct': num(ws.cell(r, 27).value),
                'marginNoAdsPct': num(ws.cell(r, 35).value),
                'marginTotalPct': num(ws.cell(r, 43).value),
                'seedTargetFillPrice': num(ws.cell(r, 68).value),
                'seedPriceRaise': num(ws.cell(r, 70).value),
                'seedTargetClientPrice': num(ws.cell(r, 72).value),
                'seedNewProfitabilityPct': num(ws.cell(r, 76).value),
                'seedProfitabilityDeltaPct': num(ws.cell(r, 79).value),
                'requiredPriceForProfitability': num(ws.cell(r, 91).value),
                'neededRaiseForProfitability': num(ws.cell(r, 94).value),
                'neededClientPriceForProfitability': num(ws.cell(r, 97).value),
                'requiredPriceForMargin': num(ws.cell(r, 104).value),
                'neededRaiseForMargin': num(ws.cell(r, 107).value),
                'neededClientPriceForMargin': num(ws.cell(r, 110).value),
            },
            'ozon': {
                'currentFillPrice': num(ws.cell(r, 58).value),
                'currentClientPrice': num(ws.cell(r, 61).value),
                'currentSppPct': num(ws.cell(r, 65).value),
                'profitabilityPct': num(ws.cell(r, 28).value),
                'marginNoAdsPct': num(ws.cell(r, 36).value),
                'marginTotalPct': num(ws.cell(r, 44).value),
                'seedTargetFillPrice': num(ws.cell(r, 69).value),
                'seedPriceRaise': num(ws.cell(r, 71).value),
                'seedTargetClientPrice': num(ws.cell(r, 73).value),
                'seedNewProfitabilityPct': num(ws.cell(r, 77).value),
                'seedProfitabilityDeltaPct': num(ws.cell(r, 80).value),
                'requiredPriceForProfitability': num(ws.cell(r, 92).value),
                'neededRaiseForProfitability': num(ws.cell(r, 95).value),
                'neededClientPriceForProfitability': num(ws.cell(r, 98).value),
                'requiredPriceForMargin': num(ws.cell(r, 105).value),
                'neededRaiseForMargin': num(ws.cell(r, 108).value),
                'neededClientPriceForMargin': num(ws.cell(r, 111).value),
            }
        }
        for market, m in market_defs.items():
            if not any(value not in (None, '') for value in (m['currentFillPrice'], m['currentClientPrice'], m['marginTotalPct'], m['profitabilityPct'])):
                continue
            row = dict(base)
            row.update({'id': f'{market}|{article_key}', 'marketplace': market})
            row.update(m)
            price_row = price_map.get((market, article_key), {})
            row.update({
                'allowedMarginPct': price_row.get('allowedMarginPct'),
                'avgMargin7dPct': price_row.get('avgMargin7dPct'),
                'minPrice': price_row.get('minPrice'),
                'basePrice': price_row.get('basePrice'),
                'monthly': price_row.get('daily', []),
                'monthlyCurrentTurnoverDays': price_row.get('currentTurnoverDays'),
                'monthlyCurrentPrice': price_row.get('currentPrice'),
            })
            rows[market].append(row)

    return {
        'generatedAt': dt.datetime.now(dt.timezone.utc).isoformat(),
        'brand': 'Алтея',
        'sourceFile': xlsx_path.name,
        'note': 'Рабочий price workbench из файла «Смарт _ Работа с ценами.xlsx». Здесь команда фиксирует решение по цене прямо в портале, а файл остаётся source/backup.',
        'platforms': {
            'wb': {'rows': rows['wb'], 'label': 'WB'},
            'ozon': {'rows': rows['ozon'], 'label': 'ОЗ'},
            'ym': {'rows': [], 'label': 'ЯМ', 'emptyNote': 'Источник по Я.Маркету ещё не подключён.'}
        }
    }


def main():
    if len(sys.argv) < 3:
        print('Usage: python build_smart_price_workbench.py <input_xlsx> <output_json> [prices_json] [skus_json]')
        raise SystemExit(1)
    xlsx_path = Path(sys.argv[1])
    output_json = Path(sys.argv[2])
    prices_json = Path(sys.argv[3]) if len(sys.argv) > 3 else None
    skus_json = Path(sys.argv[4]) if len(sys.argv) > 4 else None
    payload = build_payload(xlsx_path, prices_json, skus_json)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Saved {sum(len(v["rows"]) for v in payload["platforms"].values() if isinstance(v, dict) and "rows" in v)} rows to {output_json}')


if __name__ == '__main__':
    main()
